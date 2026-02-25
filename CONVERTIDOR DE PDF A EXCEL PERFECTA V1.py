import streamlit as st
import pandas as pd
from pdf2image import convert_from_bytes
import pytesseract
import re
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def limpiar_monto(valor):
    if not valor: return 0.0
    limpio = str(valor).replace('.', '').replace(',', '').strip()
    try: return float(limpio)
    except: return 0.0

def procesar_contable_avanzado(archivo_bytes):
    try:
        paginas = convert_from_bytes(archivo_bytes, dpi=300)
    except Exception as e:
        st.error(f"Error al leer el PDF con Poppler: {e}")
        return None

    datos_finales = []
    
    # 1. Patrón Base: Busca los extremos (Comprobante al inicio, y los 3 montos al final)
    patron_base = re.compile(r'(\S+)\s+(.*?)\s+(\d{2}-\d{2}-\d{2})\s+(\d+)\s+(.*?)\s+([-]?\s*\d[\d.,]*)\s+([-]?\s*\d[\d.,]*)\s+([-]?\s*\d[\d.,]*)$')
    
    # 2. Patrón Medio CORREGIDO: Ahora acepta comas y puntos en la Orden de Pago ([\d.,]+)
    patron_medio = re.compile(r'^(?:([\d.,]+)\s+)?(?:(\d{2}-\d{2}-\d{2})\s+)?(?:(\d+/\d{6})\s+)?(.*)$')

    progreso = st.progress(0)
    total_paginas = len(paginas)

    for i, img in enumerate(paginas):
        progreso.progress((i + 1) / total_paginas, text=f"Escaneando con OCR página {i+1} de {total_paginas}...")
        texto = pytesseract.image_to_string(img, lang='spa', config='--psm 6')
        
        for linea in texto.split('\n'):
            linea = linea.strip()
            if not linea or "Comprobante" in linea or "Totales" in linea or "Saldo inicial" in linea: 
                continue
            
            match = patron_base.search(linea)
            if match:
                medio = match.group(5).strip()
                match_m = patron_medio.match(medio)
                
                datos_finales.append({
                    "Comprobante": match.group(1),
                    "Fecha Transac.": match.group(3),
                    "Nro. Planilla": match.group(4),
                    "Tipo Planilla": "", # Se mantiene vacío intencionalmente como columna separadora
                    "Orden de Pago": match_m.group(1) if match_m.group(1) else "",
                    "Fecha de Pago": match_m.group(2) if match_m.group(2) else "",
                    "Asiento/Periodo": match_m.group(3) if match_m.group(3) else "",
                    "Descripción Concepto": match_m.group(4).strip() if match_m.group(4) else "",
                    "Débito": limpiar_monto(match.group(6)),
                    "Crédito": limpiar_monto(match.group(7)),
                    "Saldo": limpiar_monto(match.group(8))
                })

    if not datos_finales:
        return None

    df = pd.DataFrame(datos_finales)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Extracto Detallado')
        ws = writer.sheets['Extracto Detallado']

        header_fill = PatternFill(start_color="1C2C54", end_color="1C2C54", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.column_letter in ['I', 'J', 'K']: 
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        anchos = {'A': 20, 'B': 15, 'C': 12, 'D': 12, 'E': 15, 'F': 15, 'G': 18, 'H': 55, 'I': 15, 'J': 15, 'K': 15}
        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho
            
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    return output.getvalue()

# --- INTERFAZ WEB STREAMLIT ---
st.set_page_config(page_title="Extractor OCR Premium", layout="centered", page_icon="🧾")
st.title("🧾 Extractor OCR de Extractos")
st.markdown("Herramienta de extracción avanzada mediante Visión Artificial (OCR). Diseñado para procesar reportes complejos.")

archivo_subido = st.file_uploader("Sube el PDF del extracto aquí", type=["pdf"])

if archivo_subido:
    if st.button("🚀 Iniciar Extracción Avanzada"):
        with st.spinner("Inicializando motores OCR..."):
            resultado_excel = procesar_contable_avanzado(archivo_subido.read())
            
            if resultado_excel:
                st.success("✅ ¡Reporte Final Generado con éxito!")
                st.download_button(
                    label="📥 Descargar Extracto Premium",
                    data=resultado_excel,
                    file_name="Extracto_Krona_Contable_Premium.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("❌ No se detectaron filas válidas. Verifica la calidad del PDF.")
